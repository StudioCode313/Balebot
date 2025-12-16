from balethon import Client
from balethon.conditions import private, at_state
from balethon.objects import InlineKeyboard, InlineKeyboardButton
from balethon.errors.rpc_errors import ForbiddenError
from Validations import (
    validate_phone_number,
    validate_code_meli,
    validate_capacity,
    validate_price,
    validate_credit_card,
    validate_confirm
)
import os
from dotenv import load_dotenv
import json
import jdatetime
import shutil
import pandas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from persiantools.digits import fa_to_en, ar_to_fa
import re
import zipfile
import asyncio
from functools import wraps

#Variables

admin_ids = [1828929996, 238079968]
CHANNEL_ID = 4858274378
USER_LOCKS = {}
excel_file_path = 'لیست مسافران کاروان.xlsx'
signup_json_file_path = os.path.abspath(r"E:/Key_items/Karevan_BaleBot/JsonFiles/signup_datas.json")
payment_settings_json_file_path = os.path.abspath(r"E:/Key_items/Karevan_BaleBot/JsonFiles/payment_settings_datas.json")
startpanel_informations_json_file_path = os.path.abspath(r"E:/Key_items/Karevan_BaleBot/JsonFiles/startpanel_informations_datas.json")
userjoined_list_json_file_path = os.path.abspath(r"E:/Key_items/Karevan_BaleBot/JsonFiles/userjoined_list.json")
User_SignUp_Data = {}
states = {}
STATE_FILE = "states.json"
Payment_Settings_Data = []
ZIP_NAME = "photo.zip"
PASSPORT_FOLDER = "passport_photos"
RECEIPT_FOLDER = "receipt_photos"
load_dotenv()

bot = Client(os.environ["TOKEN"])
# Json Files Structures

if os.path.exists(signup_json_file_path):
    with open(signup_json_file_path, "r", encoding="utf-8") as f:
        SignUp_Datas = json.load(f)
else:
    SignUp_Datas = {
        "Name": [],
        "Phone_Number": [],
        "Code_Meli": [],
        "BirthDate": [],
        "Photo_Filepath": []
    }

SignUp_Keys = ["Name", "Phone_Number", "Code_Meli", "BirthDate", "Photo_Filepath"]

def load_states_into_ram():
    global states

    if not os.path.exists(STATE_FILE):
        states = {}
        return

    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            states = json.load(f)
    except:
        states = {} 

def get_user_lock(user_id: int) -> asyncio.Lock:
    if user_id not in USER_LOCKS:
        USER_LOCKS[user_id] = asyncio.Lock()
    return USER_LOCKS[user_id]

def user_lock_guard(func):
    @wraps(func)
    async def wrapper(*args, **kwargs):
        user_id = None

        # تشخیص user_id از message یا callback
        for arg in args:
            if hasattr(arg, "author"):
                user_id = arg.author.id
                break
            if hasattr(arg, "message") and hasattr(arg.message, "author"):
                user_id = arg.message.author.id
                break

        if user_id is None:
            # اگر user_id پیدا نشد، بدون قفل اجرا کن
            return await func(*args, **kwargs)

        lock = get_user_lock(user_id)

        async with lock:
            return await func(*args, **kwargs)

    return wrapper

def save_states_from_ram():
    global states

    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(states, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("Error writing states.json:", e)

def set_state(user_id, new_state):
    user_id = str(user_id)
    states[user_id] = new_state
    save_states_from_ram()

def get_state(user_id):
    user_id = str(user_id)
    return states.get(user_id, None)

def sanitize_filename(s: str):
    s = s.strip()
    s = re.sub(r"[\\/:\*\?\"<>\|]", "", s)
    s = re.sub(r"\s+", "_", s)
    return s


if os.path.exists(payment_settings_json_file_path):
    with open(payment_settings_json_file_path, "r", encoding="utf-8") as f:
        Payment_Settings_Datas = json.load(f)
else:
    Payment_Settings_Datas = {
        "title": "",
        "description": "",
        "price": "",
        "credit_card": ""
    }

Payment_Settings_Keys = ["title", "description", "price", "credit_card"]


if os.path.exists(startpanel_informations_json_file_path):
    with open(startpanel_informations_json_file_path, "r", encoding="utf-8") as f:
        StartPanel_Informations_Datas = json.load(f)
else:
    StartPanel_Informations_Datas = {
        "description": "",
        "signup_capacity": 0,
        "signup_count": 0,
        "trip_is_start": False
    }

#Creating or Updating json files functions

def save_signup_data_to_json():
    with open(signup_json_file_path, "w", encoding="utf-8") as f:
        json.dump(SignUp_Datas, f, ensure_ascii=False, indent=2)

save_signup_data_to_json()


def save_payment_settings_data_to_json():
    with open(payment_settings_json_file_path, "w", encoding="utf-8") as f:
        json.dump(Payment_Settings_Datas, f, ensure_ascii=False, indent=2)

save_payment_settings_data_to_json()

def save_startpanel_informations_data_to_json():
    with open(startpanel_informations_json_file_path, "w", encoding="utf-8") as f:
        json.dump(StartPanel_Informations_Datas, f, ensure_ascii=False, indent=2)

save_startpanel_informations_data_to_json()


#Checking for Payment Settings


def payment_settings_check():    
    if ("" in (Payment_Settings_Datas["title"], Payment_Settings_Datas["description"], Payment_Settings_Datas["credit_card"], Payment_Settings_Datas["price"])):
        return False
    else:
        return True


#Checking admin and membership of chanel

def is_admin(user_id):
    global admin_ids
    return user_id in admin_ids

async def check_user_membership(user_id):
    try:
        member = await bot.get_chat_member(CHANNEL_ID, user_id)
        return member.status in ("member", "creator", "administrator")

    except Exception as e:
        print(f"Error checking user membership: {e}")
        return False


#Auto shutdown

async def auto_shutdown():
    global admin_ids
    for id in admin_ids:
        try:
            await bot.send_message(id, "ثبت نام پایان یافت سفر خوبی داشته باشید.")
        except Exception as e:
            print(f"❌ Failed to send message to {id}: {e}")
            
    StartPanel_Informations_Datas["trip_is_start"] = False
    save_startpanel_informations_data_to_json()


def persian_to_english_digits(text):
    fa_digit = ar_to_fa(text)
    en_digit = fa_to_en(fa_digit)
    return en_digit


#Commands

@bot.on_command(private)
async def admin_panel(*, message):
    global StartPanel_Informations_Datas

    if is_admin(user_id= message.author.id):
        if StartPanel_Informations_Datas["trip_is_start"]:
            await message.reply(
                "پنل مدیریت",
                InlineKeyboard(
                    [("اتمام ثبت نام.", "stop_signup")],
                    [("لیست مسافران.", "passengers_list")],
                    [("تعداد نفرات باقی مانده.", "remaining_capacity")],
                    [("حذف مسافر.", "remove_passenger")],
                    [("تنظیمات پرداخت.", "payment_settings")]
                )
            )
        else:
            await message.reply(
                "پنل مدیریت",
                InlineKeyboard(
                    [("شروع ثبت نام.", "start_signup")],
                    [("تنظیمات پرداخت.", "payment_settings")],
                    [("لیست مسافران.", "passengers_list")],
                    [("حذف مسافر.", "remove_passenger")]
                )
            )
    else:
        await message.reply("شما دسترسی به این دستور را ندارید.")

    User_SignUp_Data.pop(message.author.id, None)



@bot.on_command(private) 
async def start(*, message):
    await start_core(message, message.author.id)

@user_lock_guard
async def start_core(message, user_id, client=None, *args, **kwargs):
    
    if await check_user_membership(user_id):
        await message.reply(
            StartPanel_Informations_Datas["description"],
            InlineKeyboard(
                [("ثبت نام.", "SignUp")]
            )
        )    
    else:
        await message.reply(
            "برای ادامه کار با ربات لطفا داخل چنل زیر عضو شید😊✨",
            InlineKeyboard(
                [InlineKeyboardButton('کانال کاروان', url='https://ble.ir/habib_albakin')],
                [('عضو شدم.', 'join')],
            )
        )
    remaining_capacity = StartPanel_Informations_Datas["signup_capacity"] - StartPanel_Informations_Datas["signup_count"]
    await bot.send_message(message.chat.id, f"ظریفت باقی مانده: {remaining_capacity} نفر ")
    message.author.set_state("")
    User_SignUp_Data.pop(user_id, None)


#CallBack Queryes

@bot.on_callback_query()
@user_lock_guard
async def callback_handler(callback_query, client=None, *args, **kwargs):
    global StartPanel_Informations_Datas, SignUp_Datas 
    user_id = callback_query.author.id
    set_state(callback_query.author.id, "")
    

    #Admin Panel CallBacks

    if callback_query.data == "passengers_list":
        if os.path.exists(ZIP_NAME):
            os.remove(ZIP_NAME)
        with zipfile.ZipFile(ZIP_NAME, "w", zipfile.ZIP_DEFLATED) as zipf:

    # -------- passport photos --------
            if os.path.exists(PASSPORT_FOLDER):
                for filename in os.listdir(PASSPORT_FOLDER):
                    file_path = os.path.join(PASSPORT_FOLDER, filename)
                    if os.path.isfile(file_path):
                        zipf.write(
                            file_path,
                            arcname=os.path.join("passport_photos", filename)
                        )

    # -------- receipt photos --------
            if os.path.exists(RECEIPT_FOLDER):
                for filename in os.listdir(RECEIPT_FOLDER):
                    file_path = os.path.join(RECEIPT_FOLDER, filename)
                    if os.path.isfile(file_path):
                        zipf.write(
                            file_path,
                            arcname=os.path.join("receipt_photos", filename)
                        )

        with open(signup_json_file_path, "r", encoding="utf-8") as f:
            json_SignUp_Datas = json.load(f)

        keys = list(json_SignUp_Datas.keys())
        keys_to_use = keys[:-1]

        filtered_dict = {k: json_SignUp_Datas[k] for k in keys_to_use}

        data_table = pandas.DataFrame(filtered_dict)        
        data_table.index += 1 
        data_table.columns = ['نام و نام خانوادگی', 'شماره تلفن', 'کد ملی', 'تاریخ تولد']

        data_table.to_excel(excel_file_path, index_label="ردیف")

        wb = load_workbook(excel_file_path)
        ws = wb.active

        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 30

        max_row = ws.max_row
        for row_idx in range(1, max_row + 1):
            ws.row_dimensions[row_idx].height = 60

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(size=26)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(excel_file_path)

        await bot.send_document(chat_id= callback_query.message.chat.id, document= open(excel_file_path, 'rb'))
        await bot.send_document(chat_id= callback_query.message.chat.id, document= open(ZIP_NAME, "rb"))
        await callback_query.answer("لیست مسافران در قالب فایل اکسل و عکس گذرنامه ها به صورت فشرده فرستاده شدند.")

        os.remove(ZIP_NAME)
        os.remove(excel_file_path)
        set_state(callback_query.author.id, "")

    elif callback_query.data == "remove_passenger":
        passenger_list = ""

        if StartPanel_Informations_Datas["signup_count"] > 0:
            for i, name in enumerate(SignUp_Datas["Name"]):
                passenger_list += f"{i + 1}. {name}\n"

            await callback_query.answer(f"لیست مسافران:\n\n{passenger_list}\n\nشماره مسافری که می‌خواهید حذف کنید را وارد کنید:")
            set_state(callback_query.author.id, "REMOVE_PASSENGER_SELECT")
        else:
            await callback_query.answer("هنوز مسافری ثبت نام نکرده است")
            set_state(callback_query.author.id, "")

    elif callback_query.data == "remaining_capacity":
        remaining_capacity = StartPanel_Informations_Datas["signup_capacity"] - StartPanel_Informations_Datas["signup_count"]

        await callback_query.answer(f"ظریفت باقی مانده: {remaining_capacity} نفر هست.")
        set_state(callback_query.author.id, "")

    elif callback_query.data == "payment_settings":
        await callback_query.answer("موضوع پرداخت را وارد کنید.")
        set_state(callback_query.author.id, "TITLE")

    elif callback_query.data == "start_signup":
        if (payment_settings_check()):
            await callback_query.answer("توضیحات سفر را وارد کنید.")
            set_state(callback_query.author.id, "TRIP_DESCRIPTION")

        else:
            await callback_query.answer("تنظیمات پرداخت روی هیچ مقداری تنظیم نشده است")
            set_state(callback_query.author.id, "")

    elif callback_query.data == "stop_signup":
        StartPanel_Informations_Datas["trip_is_start"] = False
        save_startpanel_informations_data_to_json()

        await callback_query.answer("ثبت نام پایان یافت سفر خوبی داشته باشید.")
        set_state(callback_query.author.id, "")

    elif callback_query.data == "confirm_receipt_no":
    # حذف فیش موقت از RAM
        try:
            User_SignUp_Data[callback_query.author.id].pop(5)  # فیش فیلد ۵ است
        except:
            pass

        set_state(callback_query.author.id, "AWAITING_RECEIPT")
        await bot.answer_callback_query(callback_query, text="لطفاً عکس فیش جدید ارسال کنید.")
        await bot.send_message(callback_query.chat.id, "لطفاً عکس فیش جدید را ارسال کنید.")

    elif callback_query.data == "confirm_receipt_yes":
        user_id = callback_query.author.id

        user_data = User_SignUp_Data.get(user_id)
        if not user_data:
            await bot.answer_callback_query(callback_query, text="خطا در اطلاعات.", show_alert=True)
            return

        name = sanitize_filename(user_data[0])
        code = sanitize_filename(str(user_data[2]))

    # مسیرها
        receipt_folder = "receipt_photos"
        passport_folder = "passport_photos"
        os.makedirs(receipt_folder, exist_ok=True)
        os.makedirs(passport_folder, exist_ok=True)

        receipt_bytes = user_data[5]
        passport_bytes = user_data[4]

    # ذخیره پاسپورت
        passport_path = f"{passport_folder}/{name}_{code}.jpg"
        with open(passport_path, "wb") as f:
            f.write(passport_bytes)
            user_data[4] = passport_path

    # ذخیره فیش
        receipt_path = f"{receipt_folder}/{name}_{code}.jpg"
        with open(receipt_path, "wb") as f:
            f.write(receipt_bytes)

    # ثبت اطلاعات کاربر داخل JSON
        for i in range(len(SignUp_Keys)):
            SignUp_Datas[SignUp_Keys[i]].append(user_data[i])

        StartPanel_Informations_Datas["signup_count"] += 1
        save_signup_data_to_json()
        save_startpanel_informations_data_to_json()

    # پیام برای مدیرها
        text = (
            f"🟢 ثبت‌نام جدید تکمیل شد.\n\n"
            f"نام: {user_data[0]}\n"
            f"شماره: {user_data[1]}\n"
            f"کدملی: {user_data[2]}\n"
            f"تولد: {user_data[3]}\n"
            f"📎 پاسپورت: {passport_path}\n"
            f"📎 فیش: {receipt_path}"
        )

        for admin in admin_ids:
            try:
                await bot.send_message(admin, text)
            except:
                pass

        await bot.send_message(callback_query.message.chat.id, "فیش تایید شد و ثبت‌نام با موفقیت به پایان رسید ✔")

        User_SignUp_Data.pop(user_id, None)
        set_state(user_id, "")

    #Start Panel CallBacks


    elif callback_query.data == "join":
        if await check_user_membership(user_id): 

            await bot.delete_message(callback_query.message.chat.id , callback_query.message.id)
            await callback_query.answer('شما عضو کانال هستید. \n حالا میتوانید برای ثبت نام اقدام کنید.')
            await start_core(callback_query.message, user_id)
            set_state(callback_query.author.id, "")

        else:
            await callback_query.answer('شما عضو کانال نیستید. \n لطفاً ابتدا عضو کانال شوید.')
            set_state(callback_query.author.id, "")

    elif callback_query.data == "SignUp":
        if (StartPanel_Informations_Datas["trip_is_start"]):
            User_SignUp_Data.pop(user_id, None)
            markup=InlineKeyboard([("لغو ثبت نام❌", "cancel_signup")])
            await bot.send_message(chat_id= callback_query.message.chat.id, text= "لطفا نام و نام خانوادگی خود را وارد نمایید✏📃", reply_markup=markup)
            set_state(callback_query.author.id, "NAME")

        else:
            await callback_query.answer("ثبت نام به پایان رسیده لطفا تا سفر بعد صبر کنید🙏💖")
            set_state(callback_query.author.id, "")

    elif callback_query.data == "cancel_signup":
        User_SignUp_Data.pop(user_id, None)

        await bot.send_message(callback_query.message.chat.id, "ثبت نام لغو شد. برای شروع مجدد /start را بزنید.")
        set_state(callback_query.author.id, "")


# remove passengers state 

@bot.on_message()
@user_lock_guard
async def stats123(message, client=None, *args, **kwargs):
    global SignUp_Datas
    global setting_payment_message_id
    if get_state(message.author.id) == "REMOVE_PASSENGER_SELECT":
        try:
            index = int(persian_to_english_digits(message.text)) - 1
            if index < 0 or index >= len(SignUp_Datas["Name"]):
                raise IndexError

            for key in SignUp_Datas.keys():
                SignUp_Datas[key].pop(index)

            StartPanel_Informations_Datas["signup_count"] -= 1
        
            save_signup_data_to_json()
            save_startpanel_informations_data_to_json()

            await message.reply("مسافر با موفقیت حذف شد.")

        except (ValueError, IndexError):
            await message.reply("شماره وارد شده معتبر نیست. \n لطفاً دوباره تلاش کنید.")

        set_state(message.author.id, "")

# Start Trip Information

    elif get_state(message.author.id) == "TRIP_DESCRIPTION":
        StartPanel_Informations_Datas["description"] = message.text

        await bot.send_message(chat_id= message.chat.id, text= "ظرفیت ثبت نام چند نفر هست؟")
        set_state(message.author.id, "SIGNUP_CAPACITY")

    elif get_state(message.author.id) == "SIGNUP_CAPACITY":
        if (validate_capacity(message.text)):

            StartPanel_Informations_Datas["signup_capacity"] = int(persian_to_english_digits(message.text))
            StartPanel_Informations_Datas["trip_is_start"] = True
            StartPanel_Informations_Datas["signup_count"] = 0
            SignUp_Datas = {
                "Name": [],
                "Phone_Number": [],
                "Code_Meli": [],
                "BirthDate": [],
                "Photo_Filepath": []
            }
            shutil.rmtree("passport_photos/")
            os.makedirs("passport_photos/")

            await bot.send_message(chat_id= message.chat.id, text= "ثبت نام با موفقیت اغاز شد.")

            save_startpanel_informations_data_to_json()
            save_signup_data_to_json()

            set_state(message.author.id, "")

        else:
            await message.reply("مقدار واد شده یک عدد معتبر نمی باشد لطفا دوباره تلاش کنید.")


       # Payment Settings

    elif get_state(message.author.id) == "TITLE":
        Payment_Settings_Data.append(message.text)
        await bot.send_message(chat_id= message.chat.id, text= "توضیحات پرداخت را وارد کنید.")
        set_state(message.author.id, "DESCRIPTION")


    elif get_state(message.author.id) == "DESCRIPTION":
        Payment_Settings_Data.append(message.text)
        await bot.send_message(chat_id= message.chat.id, text= "مبلغ را به ریال وارد کنید.")
        set_state(message.author.id, "PRICE")


    elif get_state(message.author.id) == "PRICE":
        if(validate_price(message.text)):
            Payment_Settings_Data.append(persian_to_english_digits(message.text))

            await bot.send_message(chat_id= message.chat.id, text= "شماره کارت را وارد کنید.")
            set_state(message.author.id, "CREDIT_CARD")
        else:
            await message.reply("مبلغ وارد شده معتبر نیست لطفا دوباره تلاش کنید.")

    elif get_state(message.author.id) == "CREDIT_CARD":
        if (validate_credit_card(message.text)):
            Payment_Settings_Data.append(persian_to_english_digits(message.text))

            payment_message = await bot.send_message(
                    chat_id=message.chat.id,
                    text=f'''موضوع:{Payment_Settings_Data[0]}\nتوضیحات:{Payment_Settings_Data[1]}\nقیمت:{int(Payment_Settings_Data[2])}\nشماره کارت:{int(Payment_Settings_Data[3])}''')
            setting_payment_message_id = payment_message.id

            await bot.send_message(chat_id= message.chat.id, text= "تنظیمات پرداخت را تایید میکنید؟ (بله/خیر)")
            set_state(message.author.id, "PAYMENT_CONFIRMATION")
        else:
            await message.reply("شماره کارت وارد شده معتبر نیست لطفا دوباره تلاش کنید.")


    elif get_state(message.author.id) == "PAYMENT_CONFIRMATION":
        if str(message.text).capitalize() in ("Yes", "No", "بله", "خیر"):

            if validate_confirm(message.text):

                await bot.delete_message(message.chat.id, setting_payment_message_id)

                for i in range(len(Payment_Settings_Keys)):
                    Payment_Settings_Datas[Payment_Settings_Keys[i]] = Payment_Settings_Data[i]

                Payment_Settings_Data.clear()
                save_payment_settings_data_to_json()

                await message.reply("تنظیمات پرداخت با موفقیت ثبت شد.")

                set_state(message.author.id, "")#reset state after confirmation
            else:

                await message.reply("دوباره با دستور /admin_panel تلاش کن.")
                set_state(message.author.id, "")#reset state after no confirmation
        else:
            await message.reply("لطفا دوباره تلاش کن.")


            # SignUp Process

    elif get_state(message.author.id) == "NAME":
        User_SignUp_Data[message.author.id] = [message.text]
        markup=InlineKeyboard([("لغو ثبت نام❌", "cancel_signup")])
        await bot.send_message(chat_id= message.chat.id, text= "برای ارتباط بهتر، شماره همراه خود را وارد کنید☎️📞", reply_markup=markup)
        set_state(message.author.id, "PHONE_NUMBER")

    elif get_state(message.author.id) == "PHONE_NUMBER":
        if validate_phone_number(message.text):
            User_SignUp_Data[message.author.id].append(persian_to_english_digits(message.text))
            markup=InlineKeyboard([("لغو ثبت نام❌", "cancel_signup")])
            await bot.send_message(chat_id= message.chat.id, text= "برای احراز هویت، لطفاً کد ملی معتبر خود را ثبت کنید📲🔐", reply_markup=markup)
            set_state(message.author.id, "CODE_MELI")
        else:
            await message.reply("شماره تلفن وارد شده معتبر نیست. لطفاً مجدداً بررسی و وارد نمایید📱⚠️")

    elif get_state(message.author.id) == "CODE_MELI":
        if validate_code_meli(message.text):
            User_SignUp_Data[message.author.id].append(persian_to_english_digits(message.text))
            markup=InlineKeyboard([("لغو ثبت نام❌", "cancel_signup")])
            await bot.send_message(chat_id= message.chat.id, text= "برای تکمیل اطلاعات، تاریخ تولد خود را به صورت 1364/06/15 وارد کنید🙏✨", reply_markup=markup)
            set_state(message.author.id, "BIRTHDATE")
            
        else:
            await message.reply("متأسفیم! کد ملی شما تأیید نشد. لطفاً مجدداً تلاش نمایید🙏🔄")

    elif get_state(message.author.id) == "BIRTHDATE":
        data_str = message.text
        try:
            year, month, day = map(int, data_str.split("/"))
            shamsi_date = jdatetime.date(year, month, day)
            shamsi_data = f"{shamsi_date.year}/{shamsi_date.month}/{shamsi_date.day}"
            User_SignUp_Data[message.author.id].append(shamsi_data)
            markup=InlineKeyboard([("لغو ثبت نام❌", "cancel_signup")])
            await bot.send_message(message.chat.id, "لطفاً یک تصویر واضح و خوانا از صفحه اول گذرنامه خود ارسال نمایید📸🛂", reply_markup=markup)  
            set_state(message.author.id, "PASSPORT")     
        except ValueError:
            await message.reply("تاریخ تولد وارد شده معتبر نیست. لطفاً با فرمت صحیح (مثلاً 1375/05/15) مجدداً وارد نمایید📅⚠️")

    elif get_state(message.author.id) == "PASSPORT":
        if message.photo:
            passport_photo = message.photo[-1]

            photo_file = await bot.download(passport_photo.id)
            User_SignUp_Data[message.author.id].append(photo_file)

            data = User_SignUp_Data[message.author.id]
            confirmation_message = (
                f"نام و نام‌خانوادگی: {data[0]}\n"
                f"شماره تماس: {data[1]}\n"
                f"کد ملی: {data[2]}\n "
                f"تاریخ تولد: {data[3]}\n"
                f"موارد بالا را تایید میکنید؟ (بله/خیر)"
            )
            await bot.send_message(chat_id=message.chat.id, text=confirmation_message)

            set_state(message.author.id, "SIGNUP_CONFIRMATION")

        else:
            await bot.send_message(message.chat.id, "عکس ارسال شده نامعتبر است. لطفاً تصویر واضحی از گذرنامه ارسال کنید📸🔄")

    elif get_state(message.author.id) == "SIGNUP_CONFIRMATION":
        if str(message.text).capitalize() in ("Yes", "No", "بله", "خیر"):
            if validate_confirm(message.text):
                await send_payment_text_and_wait_receipt(message)
            else:
                await message.reply("میتوانید دوباره با دستور /start ثبت نام کنید.")
                set_state(message.author.id, "")
                User_SignUp_Data.pop(message.author.id, None)
        else:
            await message.reply("متوجه نشدم, لطفا دوباره تلاش کنید.")

    elif get_state(message.author.id) == "AWAITING_RECEIPT":
        if not message.photo:
            await message.reply("لطفاً فقط *عکس فیش واریزی* ارسال کنید.")
            return

    # ذخیره موقتی عکس داخل RAM (نه داخل فایل)
        receipt_photo = message.photo[-1]
        photo_bytes = await bot.download(receipt_photo.id)

    # ذخیره داخل RAM
        User_SignUp_Data[message.author.id].append(photo_bytes)

    # استیت جدید
        set_state(message.author.id, "CONFIRM_RECEIPT")

    # درخواست تایید
        markup = InlineKeyboard([
            ("✔ تایید نهایی", "confirm_receipt_yes"),
            ("✖ ارسال دوباره", "confirm_receipt_no"),
        ])

        await bot.send_message(
            message.chat.id,
            "فیش دریافت شد.\nآیا از ارسال این فیش مطمئن هستید؟",
            reply_markup=markup
        )


async def send_payment_text_and_wait_receipt(message):
    title = Payment_Settings_Datas.get("title", "پرداخت")
    description = Payment_Settings_Datas.get("description", "")
    price = Payment_Settings_Datas.get("price", "")
    credit_card = Payment_Settings_Datas.get("credit_card", "")

    txt = (
        f"🔰 *اطلاعات پرداخت*\n\n"
        f"💠 موضوع: {title}\n"
        f"💠 توضیحات: {description}\n"
        f"💳 کارت: {credit_card}\n"
        f"💲 مبلغ: {price} ریال\n\n"
        "لطفاً هزینه را پرداخت کنید و *عکس فیش واریزی* را ارسال کنید."
    )

    await bot.send_message(message.chat.id, txt)
    set_state(message.author.id, "AWAITING_RECEIPT")


async def show_payment(message):

    try:
        user_id = int(message.successful_payment.invoice_payload)
        if user_id in User_SignUp_Data:
            User_SignUp_Data[user_id].append(True)  # mark payment as complete
    except Exception as e:
        print(f"⚠️ Payment error: {e}")
    
load_states_into_ram()
bot.run()